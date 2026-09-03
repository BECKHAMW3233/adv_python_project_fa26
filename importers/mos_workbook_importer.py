"""MOSWorkbookImporter -- discover and parse MOS worksheets."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from exceptions import SourceFileNotFoundError, WorksheetStructureError
from models import MOSCourseEquivalency, WorksheetIssue
from services.normalizer import normalize_course_id, normalize_credits, normalize_text

_MOS_CODE_PATTERN = re.compile(r"(\d{2}[A-Z])\s*[-–]?\s*(.*)")
_SKILL_LEVEL_HEADER_PATTERN = re.compile(r"Skill\s*Level\s*(\d{2})", re.IGNORECASE)
_FOOTER_MARKERS = ("Total Hours Required", "End of Worksheet", "Press UP or DOWN ARROW")
_TITLE_SEARCH_ROWS = 3
_HEADER_SEARCH_ROWS = 4


class MOSWorkbookImporter:
    """Discovers and parses every MOS worksheet in the reduced Army MOS Maps workbook."""

    def __init__(self, branch: str = "Army") -> None:
        self.branch = branch

    def import_workbook(
        self, path: Path
    ) -> tuple[list[MOSCourseEquivalency], list[WorksheetIssue]]:
        if not path.exists():
            raise SourceFileNotFoundError(f"MOS workbook not found: {path}")

        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        records: list[MOSCourseEquivalency] = []
        issues: list[WorksheetIssue] = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            try:
                records.extend(self._parse_worksheet(worksheet, sheet_name, path.name))
            except WorksheetStructureError as exc:
                issues.append(WorksheetIssue(path.name, sheet_name, str(exc)))

        return records, issues

    def _parse_worksheet(
        self, worksheet: Worksheet, sheet_name: str, source_file: str
    ) -> list[MOSCourseEquivalency]:
        rows = list(worksheet.iter_rows(values_only=True))

        mos_code, mos_title, title_row_index = self._find_mos_title(rows, sheet_name)
        header_row_index, skill_level_columns = self._find_header_row(
            rows, title_row_index, sheet_name
        )

        records: list[MOSCourseEquivalency] = []
        for row in rows[header_row_index + 1 :]:
            if self._is_footer_row(row):
                continue
            records.extend(
                self._parse_course_row(
                    row, skill_level_columns, mos_code, mos_title, source_file, sheet_name
                )
            )
        return records

    def _find_mos_title(
        self, rows: list[tuple], sheet_name: str
    ) -> tuple[str, str, int]:
        for i, row in enumerate(rows[:_TITLE_SEARCH_ROWS]):
            cell = row[0] if row else None
            if not isinstance(cell, str):
                continue
            text = cell.strip()
            if not text or any(marker in text for marker in _FOOTER_MARKERS):
                continue
            match = _MOS_CODE_PATTERN.match(text)
            if match:
                return match.group(1), normalize_text(match.group(2)), i
        raise WorksheetStructureError(
            f"Could not locate an MOS title row in the first {_TITLE_SEARCH_ROWS} rows of '{sheet_name}'"
        )

    def _find_header_row(
        self, rows: list[tuple], title_row_index: int, sheet_name: str
    ) -> tuple[int, dict[int, str]]:
        search_end = min(title_row_index + _HEADER_SEARCH_ROWS, len(rows))
        for i in range(title_row_index + 1, search_end):
            row = rows[i]
            if not row or row[0] != "Course ID":
                continue
            skill_level_columns: dict[int, str] = {}
            for col_index, header in enumerate(row):
                if not isinstance(header, str):
                    continue
                match = _SKILL_LEVEL_HEADER_PATTERN.search(header)
                if match:
                    skill_level_columns[col_index] = match.group(1)
            if not skill_level_columns:
                raise WorksheetStructureError(
                    f"Header row in '{sheet_name}' has no recognizable skill-level columns"
                )
            return i, skill_level_columns
        raise WorksheetStructureError(
            f"Could not locate a 'Course ID' header row in '{sheet_name}'"
        )

    def _is_footer_row(self, row: tuple) -> bool:
        cell = row[0] if row else None
        if not isinstance(cell, str):
            return all(c is None for c in row)
        return any(marker in cell for marker in _FOOTER_MARKERS)

    def _parse_course_row(
        self,
        row: tuple,
        skill_level_columns: dict[int, str],
        mos_code: str,
        mos_title: str,
        source_file: str,
        sheet_name: str,
    ) -> list[MOSCourseEquivalency]:
        course_id_raw = row[0]
        if not isinstance(course_id_raw, str) or not course_id_raw.strip():
            return []

        course_id = normalize_course_id(course_id_raw)
        course_title = normalize_text(row[1]) if len(row) > 1 else ""
        course_level_index = max(skill_level_columns) + 1
        course_level = (
            normalize_text(row[course_level_index])
            if len(row) > course_level_index
            else ""
        )

        records = []
        for col_index, skill_level in sorted(skill_level_columns.items()):
            raw_credits = row[col_index] if col_index < len(row) else None
            credits = normalize_credits(raw_credits)
            records.append(
                MOSCourseEquivalency(
                    mos_code=mos_code,
                    mos_title=mos_title,
                    branch=self.branch,
                    skill_level=skill_level,
                    course_id=course_id,
                    course_title=course_title,
                    credits=credits,
                    course_level=course_level,
                    source_file=source_file,
                    source_sheet=sheet_name,
                    status="ok" if credits > 0 else "no_credit",
                    notes="",
                )
            )
        return records
